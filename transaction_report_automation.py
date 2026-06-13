"""
Automated Daily Transaction Report Pipeline
Author: Diksha Mulik
Description: Automates daily banking transaction report generation
             inspired by real report automation work at Infosys.
             Reads raw CSV data, validates, cleans, and generates
             a formatted Excel report with flagged anomalies.
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
import os
import logging

# Try to import openpyxl for Excel formatting
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not installed. Run: pip install openpyxl")


# ============================================================
# CONFIGURATION
# ============================================================

CONFIG = {
    "input_file": "data/transactions.csv",
    "output_dir": "output/",
    "log_dir": "logs/",
    "report_date": date.today().strftime("%Y-%m-%d"),
    "amount_min": 1,
    "amount_max": 1000000,
    "valid_transaction_types": ["CREDIT", "DEBIT", "TRANSFER", "PAYMENT"],
}


# ============================================================
# LOGGING SETUP
# ============================================================

def setup_logging():
    """Set up logging to file and console."""
    os.makedirs(CONFIG["log_dir"], exist_ok=True)
    log_file = os.path.join(
        CONFIG["log_dir"],
        f"report_{CONFIG['report_date']}.log"
    )
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s — %(levelname)s — %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


logger = setup_logging()


# ============================================================
# STEP 1: DATA INGESTION
# ============================================================

def load_data(filepath):
    """
    Load transaction data from CSV file.
    Returns DataFrame or raises exception with clear message.
    """
    logger.info(f"Loading data from: {filepath}")
    
    if not os.path.exists(filepath):
        logger.error(f"Input file not found: {filepath}")
        raise FileNotFoundError(f"Input file not found: {filepath}")
    
    try:
        df = pd.read_csv(filepath)
        logger.info(f"Successfully loaded {len(df)} rows and {len(df.columns)} columns")
        logger.info(f"Columns: {list(df.columns)}")
        return df
    except Exception as e:
        logger.error(f"Failed to load data: {e}")
        raise


# ============================================================
# STEP 2: DATA VALIDATION
# ============================================================

def validate_data(df):
    """
    Run data quality checks on raw transaction data.
    Returns validation summary dictionary.
    """
    logger.info("Running data validation checks...")
    
    validation_results = {
        "total_records": len(df),
        "issues_found": [],
        "flagged_records": pd.DataFrame()
    }
    
    flagged_list = []

    # Check 1: Null values in critical fields
    critical_fields = ["transaction_id", "account_id", "amount", "transaction_date"]
    for field in critical_fields:
        if field in df.columns:
            null_count = df[field].isnull().sum()
            if null_count > 0:
                issue = f"NULL values in {field}: {null_count} records"
                validation_results["issues_found"].append(issue)
                logger.warning(issue)
                null_rows = df[df[field].isnull()].copy()
                null_rows["flag_reason"] = f"NULL_{field.upper()}"
                flagged_list.append(null_rows)

    # Check 2: Duplicate transaction IDs
    if "transaction_id" in df.columns:
        dupes = df[df.duplicated(subset=["transaction_id"], keep=False)]
        if len(dupes) > 0:
            issue = f"Duplicate transaction IDs: {len(dupes)} records"
            validation_results["issues_found"].append(issue)
            logger.warning(issue)
            dupes_copy = dupes.copy()
            dupes_copy["flag_reason"] = "DUPLICATE_TRANSACTION_ID"
            flagged_list.append(dupes_copy)

    # Check 3: Invalid amount values
    if "amount" in df.columns:
        invalid_amount = df[
            (df["amount"] < CONFIG["amount_min"]) | 
            (df["amount"] > CONFIG["amount_max"])
        ]
        if len(invalid_amount) > 0:
            issue = f"Invalid amount values: {len(invalid_amount)} records"
            validation_results["issues_found"].append(issue)
            logger.warning(issue)
            invalid_copy = invalid_amount.copy()
            invalid_copy["flag_reason"] = "INVALID_AMOUNT"
            flagged_list.append(invalid_copy)

    # Check 4: Invalid transaction types
    if "transaction_type" in df.columns:
        invalid_types = df[
            ~df["transaction_type"].str.upper().isin(CONFIG["valid_transaction_types"])
        ]
        if len(invalid_types) > 0:
            issue = f"Invalid transaction types: {len(invalid_types)} records"
            validation_results["issues_found"].append(issue)
            logger.warning(issue)
            invalid_copy = invalid_types.copy()
            invalid_copy["flag_reason"] = "INVALID_TRANSACTION_TYPE"
            flagged_list.append(invalid_copy)

    # Combine all flagged records
    if flagged_list:
        validation_results["flagged_records"] = pd.concat(flagged_list).drop_duplicates()

    total_issues = len(validation_results["issues_found"])
    logger.info(f"Validation complete. {total_issues} issue types found.")
    logger.info(f"Total flagged records: {len(validation_results['flagged_records'])}")

    return validation_results


# ============================================================
# STEP 3: DATA CLEANING
# ============================================================

def clean_data(df):
    """
    Clean and standardize transaction data.
    Returns cleaned DataFrame.
    """
    logger.info("Starting data cleaning...")
    original_count = len(df)
    
    df_clean = df.copy()

    # Standardize column names
    df_clean.columns = df_clean.columns.str.strip().str.lower().str.replace(" ", "_")

    # Parse and standardize date column
    if "transaction_date" in df_clean.columns:
        df_clean["transaction_date"] = pd.to_datetime(
            df_clean["transaction_date"], errors="coerce"
        )
        invalid_dates = df_clean["transaction_date"].isnull().sum()
        if invalid_dates > 0:
            logger.warning(f"Removed {invalid_dates} rows with unparseable dates")
        df_clean = df_clean.dropna(subset=["transaction_date"])

    # Standardize transaction type to uppercase
    if "transaction_type" in df_clean.columns:
        df_clean["transaction_type"] = df_clean["transaction_type"].str.strip().str.upper()

    # Remove leading/trailing whitespace from string columns
    string_cols = df_clean.select_dtypes(include="object").columns
    for col in string_cols:
        df_clean[col] = df_clean[col].str.strip()

    # Remove rows with null critical fields
    critical = [c for c in ["transaction_id", "account_id", "amount"] if c in df_clean.columns]
    df_clean = df_clean.dropna(subset=critical)

    # Remove exact duplicates
    df_clean = df_clean.drop_duplicates()

    # Ensure amount is numeric
    if "amount" in df_clean.columns:
        df_clean["amount"] = pd.to_numeric(df_clean["amount"], errors="coerce")
        df_clean = df_clean.dropna(subset=["amount"])

    cleaned_count = len(df_clean)
    removed = original_count - cleaned_count
    logger.info(f"Cleaning complete. {removed} rows removed. {cleaned_count} clean records remaining.")

    return df_clean


# ============================================================
# STEP 4: DATA TRANSFORMATION & SUMMARY
# ============================================================

def generate_summary(df):
    """
    Generate transaction summary statistics for the report.
    Returns dictionary of summary DataFrames.
    """
    logger.info("Generating summary statistics...")

    summaries = {}

    # Overall summary
    summaries["overall"] = pd.DataFrame({
        "Metric": [
            "Total Transactions",
            "Total Amount",
            "Average Amount",
            "Max Amount",
            "Min Amount",
            "Unique Accounts"
        ],
        "Value": [
            len(df),
            f"₹{df['amount'].sum():,.2f}" if "amount" in df.columns else "N/A",
            f"₹{df['amount'].mean():,.2f}" if "amount" in df.columns else "N/A",
            f"₹{df['amount'].max():,.2f}" if "amount" in df.columns else "N/A",
            f"₹{df['amount'].min():,.2f}" if "amount" in df.columns else "N/A",
            df["account_id"].nunique() if "account_id" in df.columns else "N/A"
        ]
    })

    # Summary by transaction type
    if "transaction_type" in df.columns and "amount" in df.columns:
        summaries["by_type"] = df.groupby("transaction_type").agg(
            Total_Transactions=("transaction_id", "count"),
            Total_Amount=("amount", "sum"),
            Average_Amount=("amount", "mean"),
            Max_Amount=("amount", "max")
        ).round(2).reset_index()

    # Summary by date
    if "transaction_date" in df.columns and "amount" in df.columns:
        df["date_only"] = df["transaction_date"].dt.date
        summaries["by_date"] = df.groupby("date_only").agg(
            Total_Transactions=("transaction_id", "count"),
            Total_Amount=("amount", "sum"),
            Average_Amount=("amount", "mean")
        ).round(2).reset_index()
        summaries["by_date"].columns = ["Date", "Total_Transactions", "Total_Amount", "Average_Amount"]

    logger.info("Summary generation complete.")
    return summaries


# ============================================================
# STEP 5: EXCEL REPORT GENERATION
# ============================================================

def generate_excel_report(df_clean, summaries, validation_results):
    """
    Generate formatted Excel report with multiple sheets.
    """
    os.makedirs(CONFIG["output_dir"], exist_ok=True)
    output_file = os.path.join(
        CONFIG["output_dir"],
        f"transaction_report_{CONFIG['report_date']}.xlsx"
    )

    logger.info(f"Generating Excel report: {output_file}")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # Sheet 1: Overall Summary
        summaries["overall"].to_excel(writer, sheet_name="Summary", index=False)

        # Sheet 2: By Transaction Type
        if "by_type" in summaries:
            summaries["by_type"].to_excel(writer, sheet_name="By Type", index=False)

        # Sheet 3: By Date
        if "by_date" in summaries:
            summaries["by_date"].to_excel(writer, sheet_name="By Date", index=False)

        # Sheet 4: Clean Transaction Data
        df_clean.to_excel(writer, sheet_name="All Transactions", index=False)

        # Sheet 5: Flagged Records
        if len(validation_results["flagged_records"]) > 0:
            validation_results["flagged_records"].to_excel(
                writer, sheet_name="Flagged Records", index=False
            )
        else:
            pd.DataFrame({"Message": ["No flagged records found. All data passed validation."]}).to_excel(
                writer, sheet_name="Flagged Records", index=False
            )

        # Sheet 6: Validation Log
        issues = validation_results["issues_found"]
        if issues:
            pd.DataFrame({"Validation Issues": issues}).to_excel(
                writer, sheet_name="Validation Log", index=False
            )
        else:
            pd.DataFrame({"Validation Issues": ["All checks passed. No issues found."]}).to_excel(
                writer, sheet_name="Validation Log", index=False
            )

    logger.info(f"Excel report saved: {output_file}")
    return output_file


# ============================================================
# STEP 6: APPLY EXCEL FORMATTING
# ============================================================

def apply_formatting(output_file):
    """
    Apply professional formatting to the Excel report.
    - Header row: blue background, white bold text
    - Flagged records: red highlight
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Skipping formatting.")
        return

    logger.info("Applying Excel formatting...")

    wb = load_workbook(output_file)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    flag_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        # Highlight flagged records sheet in red
        if sheet_name == "Flagged Records":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = flag_fill

    wb.save(output_file)
    logger.info("Formatting applied successfully.")


# ============================================================
# MAIN PIPELINE
# ============================================================

def run_pipeline():
    """
    Main function that runs the full report generation pipeline.
    Mimics the automated daily report job that runs on schedule.
    """
    logger.info("=" * 60)
    logger.info(f"DAILY TRANSACTION REPORT PIPELINE STARTED")
    logger.info(f"Report Date: {CONFIG['report_date']}")
    logger.info("=" * 60)

    try:
        # Step 1: Load data
        df_raw = load_data(CONFIG["input_file"])

        # Step 2: Validate
        validation_results = validate_data(df_raw)

        # Step 3: Clean
        df_clean = clean_data(df_raw)

        # Step 4: Generate summaries
        summaries = generate_summary(df_clean)

        # Step 5: Generate Excel report
        output_file = generate_excel_report(df_clean, summaries, validation_results)

        # Step 6: Apply formatting
        if OPENPYXL_AVAILABLE:
            apply_formatting(output_file)

        logger.info("=" * 60)
        logger.info("PIPELINE COMPLETED SUCCESSFULLY")
        logger.info(f"Report saved to: {output_file}")
        logger.info(f"Total records processed: {len(df_clean)}")
        logger.info(f"Flagged records: {len(validation_results['flagged_records'])}")
        logger.info("=" * 60)

        return output_file

    except Exception as e:
        logger.error(f"PIPELINE FAILED: {e}")
        raise


if __name__ == "__main__":
    run_pipeline()
